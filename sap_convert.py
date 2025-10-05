#
#   TagBotFT - Tag Bot For Tables
#   Convert SAP data
#

import os
import platform
import pandas as pd
import sys
import numpy as np
import time
import datetime
import configparser
import glob
import re
from itertools import repeat

# Standard printout of TagBotFT messages
def message(msg_text):
    print()
    print("-"*78)
    print(msg_text)
    print("-"*78)
    return

def convertAmount(amount: str) -> float:
    temp = amount.strip().replace('.', '').replace(',', '.')
    return float(temp) if temp[-1] != '-' else \
    float('-' + temp[:-1])

# Read the excludes file into dataframe
def readExclude(filename):
    message("Reading exclude list")

    try:
        # Pandas CSV function takes the first line as column name
        df = pd.read_csv(filename, delimiter=' ')
        print("Columns with exclude data:", list(df.columns))
        return df

    # If there is no file return an empty data frame
    except FileNotFoundError:
        print("No exclude file")
        return pd.DataFrame()

# Keep only aligned rows in string columns
def maxcol(newData, work_data):

    message('Removing not aligned entries')

    # Get the total number of rows at start
    org_length = len(newData)

    # Iterate over all the existing columns
    for col_name in work_data.columns:
        # Only iterate if the column is a string column
        if work_data[col_name].dtype == object:
            # Get the length of the column
            col_length = work_data[col_name].str.len()
            # Get the data where there is only one length
            if len(col_length.groupby(col_length)) == 1:
                print(col_name, len(col_length.groupby(col_length)))
                # Keep the matching rows according
                newData = newData[newData[col_name].str.match('\w{'\
                        +str(col_length[0])+'}$')]
    print('Remaining input data entries:', len(newData), 'of', org_length)
    return newData

# Dickmann / Birkenkamp
# Read SAP data files with RegEx to identify relevant data lines and items
# Input files can come from Kosten, GWGs oder Anlagen
def read_SAP_1(files: str, exclude_file, work_data):

    # Reading exclude file
    exclude_df = readExclude(exclude_file)

    message("Loading SAP data files")

    # Extract unclassified data from possibly several files, ignoring
    # data rows with cost centres contained in a list of cost centres to
    # be excluded.
    SAPinput = []
    rownum = 0

    # Get the list of files to process
    file_list = list(glob.glob(files))

    for f in file_list:
        try:
            rf = open(f, encoding='latin-1')
        except FileNotFoundError:
            print('File ' + f + ' not found.')
        else:
            # Counter for all lines
            totalLineCounter = 0
            lineCounter = 0
            # Regular expression pattern for cost files.
            pc = re.compile(
                r'\|\s+(\d{4})'                # Jahr.
                 '\|(\d+\s+)'                  # Kostenstelle.
                 '\|(\d+\s+)'                  # Kostenart.
                 '\|[^\|]+'                    # Kostenartenbezeichn.
                 '\|(\d{2}\.\d{2}\.\d{4})\s+'  # Buchungsdatum.
                 '\|\s*([\d\.]+,\d{2}[\s-])'   # Wert/KWähr.
                 '\|([^\|]+)'                  # Bezeichnung.
                 '\|([^\|]+)'                  # Bestelltext.
                 '\|[^\|]+'                    # Partnerobjektbezeichnung.
                 '\|\s+'                       # Partner-Kstl.
                 '\|(\d+|)\s+'                 # Einkaufsbeleg.
                 '\|\d*\s*'                    # Konto Gegenbuchung.
                 '\|([^\|]+)')               # Ref Beleg.
            # Regular expression pattern for asset files.
            pa = re.compile(
                r'\|([^\|]+)'                  # Kostenst.
                 '\|\d+\s+'                    # Finanzst.
                 '\|\d+\s*'                    # UNr.
                 '\|(\d{2}\.\d{2}\.(\d{4}))'   # Aktivdatum.
                 '\|([^\|]+)'                  # Anlagenbezeichnung.
                 '\|\s*([\d\.]+,\d{2})\s'      # AnschWert.
                 '\|[^\|]+'                    # Währg.
                 '\|[^\|]+'                    # Anlagenbezeichnung.
                 '\|(\d{7})[\s\d]'             # Anlage.
                 '\|\d+\s+\|')                 # Geldg.
            # Regular expression pattern for alternative asset files.
            paa = re.compile(
                r'\|(\d{7})\s*'                # Anlage.
                 '\|\d*\s*'                    # UNr.
                 '\|(\d{7})\s*'                # Kostenst.
                 '\|\d*\s*'                    # Finanzst.
                 '\|(\d{2}\.\d{2}\.(\d{4}))'   # Aktivdatum.
                 '\|([^\|]+)'                  # Anlagenbezeichnung.
                 '\|\s*([\d\.]+,\d{2})\s'      # AnschWert.
                 '\|[^\|]+'                    # Währg.
                 '\|[^\|]+'                    # Anlagenbezeichnung.
                 '\|\d*\s*\|')                 # Geldg.
            # Regular expression pattern for small asset files.
            pg = re.compile(
                r'\|[^\|]+'                    # Kostenst.
                 '\|(\d+\s+)'                  # Finanzst.
                 '\|[^\|]+'                    # Fistl-Text
                 '\|\d+\s+'                    # Finanzpos.
                 '\|[^\|]+'                    # Vorg. Nr.
                 '\|(\d{2}\.\d{2}\.(\d{4}))'   # HHM-Budat
                 '\|\s*([\d\.]+,\d{2}[\s-])'   # Zahl.Budg.
                 '\|[^\|]+'                    # Kreditor
                 '\|([^\|]+)'                  # Text
                 '\|([^\|]+)'                  # Kurztext
                 '\|([^\|]+)'                  # Bezeichg
                 '\|(\d+\s+|\d+|\s+)'          # RefBelegnr
                 '\|\s+\d+'                    # Pos.
                 '\|\d+\s+'                    # Btr.art
                 '\|[^\|]+'                    # Betragsart
                 '\|[^\|]+'                    # Fipos-Text
                 '\|([^\|]+)')                 # RW-Beleg
            for line in rf:
                totalLineCounter += 1
                # Check for match of Kosten data file
                mc = pc.match(line)
                if mc:
                    cost_cntr = mc.group(2).rstrip()
                    if cost_cntr.isdigit():
                        SAPinput.append((
                            'Kosten',
                            cost_cntr.strip(),                # Kostenart
                            (mc.group(6).strip() + ' ' + \
                             mc.group(7).strip()).upper(),       # Text.
                            mc.group(3).strip(),
                            convertAmount(mc.group(5)),        # Amount.
                            mc.group(4),                          # Date.
                            mc.group(1).strip(),                  # Year.
                            mc.group(8).strip()))                 # Ref. num.
                        rownum += 1
                        lineCounter += 1

                # Check for match of Anlagen data file
                ma = pa.match(line)
                if ma:
                    cost_cntr = ma.group(1).rstrip()
                    if cost_cntr.isdigit():
                        SAPinput.append((
                            'Anlagen',
                            cost_cntr.strip(),
                            ma.group(4).strip().upper().strip(),
                            '', convertAmount(ma.group(5)),        # Amount.
                            mc.group(2),                          # Date.
                            ma.group(3).strip(),                  # Year.
                            ma.group(2).strip()))                 # Anl. num.
                        rownum += 1
                        lineCounter += 1

                # Check for match of alternative Anlagen data file
                ma = paa.match(line)
                if ma:
                    cost_cntr = ma.group(2).rstrip()
                    if cost_cntr.isdigit():
                        SAPinput.append((
                            'Anlagen',
                            cost_cntr.strip(),
                            ma.group(5).strip().upper().strip(),
                            '', convertAmount(ma.group(6)),        # Amount.
                            ma.group(3),                          # Date.
                            ma.group(4).strip(),                  # Year.
                            ma.group(1).strip()))                 # Ref. num.
                        rownum += 1
                        lineCounter += 1

                # Check for match of GWG data file
                ma = pg.match(line)
                if ma:
                    cost_cntr = ma.group(1).rstrip()
                    if cost_cntr.isdigit():
                        SAPinput.append((
                            'Anlagen',
                            cost_cntr.strip(),
                            ma.group(5).strip().upper().strip(),
                            '', convertAmount(ma.group(4)),        # Amount.
                            ma.group(2),                          # Date.
                            ma.group(3).strip(),                  # Year.
                            ma.group(8).strip()))                 # Ref. num.
                        rownum += 1
                        lineCounter += 1
            rf.close()
            print("Read " + str(lineCounter) + " of " + \
                    str(totalLineCounter) + " lines from " + str(f))

    # Just get a limited number of items in test mode
    SAPinput = SAPinput

    #print(SAPinput)
    # Generate a dataframe from input
    df = pd.DataFrame(data=SAPinput, \
                     columns=['Art', 'Kostenstelle', \
                     'Text', 'Kostenart', 'Betrag', 'Datum', 'Jahr', \
                     'Referenz'])

    df['Datum'] = pd.to_datetime(df['Datum'], format='%d.%m.%Y')
    df = df.replace(r"^None$", '', regex=True)

    # Filtering the the SAP data by the exclude data
    if len(exclude_df) > 1:
        newData = filter_input(df, exclude_df)
    else:
        newData = df

    # Remove uneven string columns
    if len(work_data) > 1:
        newData = maxcol(newData, work_data)

    return newData

# Writing the results to an Excel file
def writeXLS(filename_w, results):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import NamedStyle
    from openpyxl.styles import Font, Fill

    wb = Workbook()
    sheet = wb.active

    a = 1
    for col_name in results.columns:
        if results[col_name].dtype == 'datetime64[ns]':
            sheet.column_dimensions[get_column_letter(a)].width = 10
        sheet.cell(row=1, column=a).value = col_name
        sheet.cell(row=1, column=a).font = Font(name='Arial', size=10)
        a += 1

    # Start in row 2 after the header row
    i = 2
    for col_name, wd in results.iterrows():

        # Assign the rest in a loop
        a = 1
        for cell_value in wd:
            if type(cell_value) == pd._libs.tslibs.timestamps.Timestamp:
                cell_value = cell_value.date()
            sheet.cell(row=i, column=a).value = cell_value
            sheet.cell(row=i, column=a).font = Font(name='Calibri', size=10)
            a += 1

        if i > 1 and type(sheet['M'+str(i)].value) == float:
            # Color the line until quality value with color
            # First get the value and set the color defaults
            red = 'efbdbd'
            yellow = 'efeebd'
            val = float(sheet['M'+str(i)].value)

            if val < 0.7:
                color_row(sheet, 'A'+str(i)+':'+'M'+str(i), red)
            if val >= 0.7 and val < 0.9:
                color_row(sheet, 'A'+str(i)+':'+'M'+str(i), yellow)
        i += 1

    # Add Autofilter
    sheet.auto_filter.ref = "A1:BA" + str(i)

    # Set auto adjusted column widths
    for idx, col in enumerate(sheet.columns, 1):
        sheet.column_dimensions[get_column_letter(idx)].auto_size = True

    # Fixate the top row
    c = sheet['A2']
    sheet.freeze_panes = c

    # Don't forget to save the file
    wb.save(filename = filename_w)
    return None

#
# Convertig data starts
#
if __name__ == '__main__':

    message('\nStarting SAP data conversion\n')

    # Setting initial parameters
    # Set default variables
    ini_file = 'sap_convert.ini'
    test = False
    myDir = os.path.abspath(os.path.dirname(__file__)) + '/'
    config = configparser.ConfigParser()
    config.sections()
    config.read(myDir + ini_file)
    input_files = config['Settings']['input_files']
    input_dir = config['Settings']['input_dir']
    input_data = input_dir + "/" + input_files
    output_dir = config['Settings']['output_dir']

    from pathlib import Path
    home = str(Path.home())

    input_data = input_data.replace("~", home)
    output_dir = output_dir.replace("~", home) + "/"

    # Read the input data into dataframe
    convData = read_SAP_1(input_data, '', '')

    # Writing tagging results to Excel file
    writeXLS(output_dir + 'sap_result_conv.xlsx', convData)
    convData.to_csv(output_dir + 'sap_result_conv.csv', sep=";")

    # End of TagBotFT
    message('SAP data conversion has finished')
    