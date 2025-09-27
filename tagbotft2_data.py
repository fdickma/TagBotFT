import pandas as pd
import sys
import time
import pandas as pd
import re
import os
import glob
from sqlalchemy import create_engine
from sqlalchemy import create_engine, MetaData
import __main__

# Function to save data to a database in a certain table
def save_data(to_store, database_name, table_name):
    print("Saving to database:\t", table_name)

    # Define database connection
    engine = create_engine('sqlite:///' + database_name, echo=False)
    sqlite_connection = engine.connect()

    # Save data to database
    to_store.to_sql(name=table_name, con=engine, if_exists='replace')

    # Close database connection
    sqlite_connection.close()

# Function to read data from a database from a certain table
def read_data(database_name, table_name):
    print("Reading from database:\t", table_name)

    # Define database connection
    engine = create_engine('sqlite:///' + database_name, echo=False)
    sqlite_connection = engine.connect()

    # Read data from database
    temp_df = pd.read_sql('SELECT * FROM ' + table_name, con=engine, index_col='index')
    
    # Close database connection
    sqlite_connection.close()
    return temp_df

def file_read(file_path):

    # Get the list of files to process
    file_list = list(glob.glob(file_path))

    # Define temporary Dataframe
    read_df = pd.DataFrame()

    if len(file_list) < 1:
        return read_df

    # Iterate the given file list
    for f in file_list:
        
        if __main__.max_lines > 0 and len(read_df) >= __main__.max_lines:
            continue
        
        print("Reading file data:\t", f)
        
        try:
            # Pandas CSV function takes the first line as column name
            if f[-4:] == '.csv':
                if __main__.max_lines > 0:
                    read_df = pd.read_csv(f, delimiter=';', nrows=__main__.max_lines)
                else:
                    read_df = pd.read_csv(f, delimiter=';')

            # Read the Excel file into a DataFrame
            # If max lines is set, read only the number of max lines from the file
            if f[-5:] == '.xlsx':
                if __main__.max_lines > 0:
                    read_df = pd.read_excel(f, nrows=__main__.max_lines)
                # Otherwise load the complete file
                else:
                    read_df = pd.read_excel(f)

            if len(read_df) > 0:
                try:
                    file_data = pd.concat([read_df, q], ignore_index=True)
                except:
                    file_data = read_df

        except FileNotFoundError:
            print('File ' + f + ' not found.')
    
    return file_data

# Check database if a certain table exists
def check_table(database_name, table_name):
    
    # Check for the database first, return False if it does not exist
    try:
        engine = create_engine('sqlite:///' + database_name, echo=False)
    except:
        return False

    metadata = MetaData()
    metadata.reflect(bind=engine)
    try: 
        my_table = metadata.tables[table_name]
    except:
        my_table = None

    if my_table is None:
        print(table_name, "does not exist")
        return False
    else:
        return True

# Coloring a set of cells in a row
def color_row(sheet, cell_range, color):
    from openpyxl.styles import PatternFill
    
    fill = PatternFill("solid", fgColor = color)
    
    # Iterate through the cell range
    rows = sheet[cell_range]
    for row in rows:
        
        # Iterate through the cells
        for c in row:
            c.fill = fill

# Writing the results to an Excel file
def write_results(result_data):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import NamedStyle
    from openpyxl.styles import Font, Fill

    wb = Workbook()
    sheet = wb.active
    
    a = 1            
    for col_name in result_data.columns:
        sheet.column_dimensions[get_column_letter(a)].width = \
            result_data[col_name].astype(str).str.len().max()
        if col_name == "TB_qual":
            sheet.column_dimensions[get_column_letter(a)].width = 12
            quality_col = get_column_letter(a)
        sheet.cell(row=1, column=a).value = col_name
        sheet.cell(row=1, column=a).font = Font(name='Arial', size=10)
        a += 1

    # Start in row 2 after the header row
    i = 2
    for col_name, wd in result_data.iterrows():
        
        # Assign the rest in a loop
        a = 1
        for cell_value in wd:
            if type(cell_value) == pd._libs.tslibs.timestamps.Timestamp:
                cell_value = cell_value.date()
            sheet.cell(row=i, column=a).value = cell_value
            sheet.cell(row=i, column=a).font = Font(name='Calibri', size=10)
            a += 1

        if i > 1 and (type(sheet[quality_col+str(i)].value) == float or \
            type(sheet[quality_col+str(i)].value) == int):
            # Color the line until quality value with color
            # First get the value and set the color defaults
            red = 'efbdbd'
            yellow = 'efeebd'
            val = float(sheet[quality_col+str(i)].value)

            if val > 1:
                val = val / 100
            
            if val < 0.2:
                color_row(sheet, 'A'+str(i)+':'+quality_col+str(i), red)
            if val >= 0.2 and val < 0.9:
                color_row(sheet, 'A'+str(i)+':'+quality_col+str(i), yellow)
        i += 1
            
    # Add Autofilter
    sheet.auto_filter.ref = "A1:BA" + str(i)

    # Set auto adjusted column widths
    for idx, col in enumerate(sheet.columns, 1):
        sheet.column_dimensions[get_column_letter(idx)].auto_size = True

    # Fixate the top row
    c = sheet['A2']
    sheet.freeze_panes = c

    # Don't forget to save the file
    wb.save(filename = "results.xlsx")
    print("Results written:\t", i - 2)
    return None
