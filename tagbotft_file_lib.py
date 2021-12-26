import re
import glob
import time
import datetime
import pandas as pd
import numpy as np

import tagbotft_lib as tl

# Read input data from excel file and a defined worksheet
def read_xl_learn(learn_file, wsheet, max_rows, max_cols, tag_col, text_col):
    from openpyxl import load_workbook
    # Loading data file with a maximum number of lines
    print("Reading Excel sheet")
    wb = load_workbook(learn_file, data_only=True, read_only=True)
    sheet = wb[wsheet]
    
    # Get the real maximum of lines to read
    if sheet.max_row < max_rows:
        max_rows = sheet.max_row

    org_data = []
    # Start with no lines read
    line_count = 0
    for row in sheet.rows:
        wsrow = []
        row_i = 0
        for cell in row:
            if row_i < max_cols:
                wsrow.append(cell.value)
                row_i += 1
        if line_count > max_rows:
            break
        else:
            org_data.append(wsrow)
            # Calculating the progress as percentage
            progress = line_count / max_rows * 100
            line_count += 1
            print("Reading file: %.2f %%" % progress, end='\r', flush=True)
    print(f'\n\rLoaded {line_count} lines.')
    print("Headers:", org_data[0])

    # Arrange the Excel data in a Pandas dataframe
    dfs = pd.DataFrame(data=org_data[1:], columns=org_data[0])

    # Setting default colums in dataframe
    dfs = dfs.rename(columns={tag_col: 'Tag', text_col: 'Text'})

    # Ensure consistent column data for columns of string value
    for sheet_col in dfs:
        if dfs[sheet_col].dtypes == object:
            dfs[sheet_col] = dfs[sheet_col].astype(str)
        
    print(dfs)
    return dfs

# Write log file
def writeLog(logEvent):
    f = open('tagbotft.log', 'a')
    dt = datetime.datetime.now()
    stamp = dt.strftime("%Y-%m-%d %H:%M:%S")
    print(logEvent)
    f.write(str(stamp) + " : " + logEvent + "\n")
    f.close()
    
# Dickmann / Birkenkamp
# Read SAP data files with RegEx to identify relevant data lines and items
# Input files can come from Kosten, GWGs oder Anlagen
def read_SAP_1(files: str, \
               test_len: int, test: bool) -> list:
    # Extract unclassified data from possibly several files, ignoring
    # data rows with cost centres contained in a list of cost centres to
    # be excluded.
    SAPinput = []
    rownum = 0
    
    # Get the list of files to process
    file_list = list(glob.glob(files))
    
    for f in file_list:
        try:
            rf = open(f, encoding='latin-1')
        except FileNotFoundError:
            print('File ' + f + ' not found.')
        else:
            # Counter for all lines
            totalLineCounter = 0
            lineCounter = 0
            # Regular expression pattern for cost files.
            pc = re.compile(
                r'\|(\d{4})\s+'                # Jahr.
                 '\|(\d+\s+)'                  # Kostenstelle.
                 '\|(\d+\s+)'                  # Kostenart.
                 '\|[^\|]+'                    # Kostenartenbezeichn.
                 '\|(\d{2}\.\d{2}\.\d{4})\s+'  # Buchungsdatum.
                 '\|\s*([\d\.]+,\d{2}[\s-])'   # Wert/KWähr.
                 '\|([^\|]+)'                  # Bezeichnung.
                 '\|([^\|]+)'                  # Bestelltext.
                 '\|[^\|]+'                    # Partnerobjektbezeichnung.
                 '\|\s+'                       # Partner-Kstl.
                 '\|(\d+|)\s+'                 # Einkaufsbeleg.
                 '\|\d*\s*'                    # Konto Gegenbuchung.
                 '\|([^\|]+)')               # Ref Beleg.
            # Regular expression pattern for asset files.
            pa = re.compile(
                r'\|([^\|]+)'                  # Kostenst.
                 '\|\d+\s+'                    # Finanzst.
                 '\|\d+\s*'                    # UNr.
                 '\|(\d{2}\.\d{2}\.(\d{4}))'   # Aktivdatum.
                 '\|([^\|]+)'                  # Anlagenbezeichnung.
                 '\|\s*([\d\.]+,\d{2})\s'      # AnschWert.
                 '\|[^\|]+'                    # Währg.
                 '\|[^\|]+'                    # Anlagenbezeichnung.
                 '\|(\d{7})[\s\d]'             # Anlage.
                 '\|\d+\s+\|')                 # Geldg.
            # Regular expression pattern for alternative asset files.
            paa = re.compile(
                r'\|(\d{7})\s*'                # Anlage.
                 '\|\d*\s*'                    # UNr.
                 '\|(\d{7})\s*'                # Kostenst.
                 '\|\d*\s*'                    # Finanzst.
                 '\|(\d{2}\.\d{2}\.(\d{4}))'   # Aktivdatum.
                 '\|([^\|]+)'                  # Anlagenbezeichnung.
                 '\|\s*([\d\.]+,\d{2})\s'      # AnschWert.
                 '\|[^\|]+'                    # Währg.
                 '\|[^\|]+'                    # Anlagenbezeichnung.
                 '\|\d*\s*\|')                 # Geldg.
            # Regular expression pattern for small asset files.
            pg = re.compile(
                r'\|[^\|]+'                    # Kostenst.
                 '\|(\d+\s+)'                  # Finanzst.
                 '\|[^\|]+'                    # Fistl-Text 
                 '\|\d+\s+'                    # Finanzpos.
                 '\|[^\|]+'                    # Vorg. Nr.
                 '\|(\d{2}\.\d{2}\.(\d{4}))'   # HHM-Budat
                 '\|\s*([\d\.]+,\d{2}[\s-])'   # Zahl.Budg.
                 '\|[^\|]+'                    # Kreditor
                 '\|([^\|]+)'                  # Text
                 '\|([^\|]+)'                  # Kurztext
                 '\|([^\|]+)'                  # Bezeichg
                 '\|(\d+\s+|\d+|\s+)'          # RefBelegnr
                 '\|\s+\d+'                    # Pos.
                 '\|\d+\s+'                    # Btr.art
                 '\|[^\|]+'                    # Betragsart
                 '\|[^\|]+'                    # Fipos-Text
                 '\|([^\|]+)')                 # RW-Beleg
            for line in rf:
                totalLineCounter += 1
                # Check for match of Kosten data file
                mc = pc.match(line)
                if mc:
                    cost_cntr = mc.group(2).rstrip()
                    if cost_cntr.isdigit():
                        SAPinput.append((
                            'Kosten', 
                            cost_cntr.strip(), '',                # Kostenart 
                            (mc.group(6).strip() + ' ' + \
                             mc.group(7).strip()).upper(),       # Text.
                            mc.group(3).strip(),
                            tl.convertAmount(mc.group(5)),        # Amount.
                            mc.group(4),                          # Date.
                            mc.group(1).strip(),                  # Year.
                            '', '',
                            mc.group(8).strip()))                 # Ref. num.
                        rownum += 1
                        lineCounter += 1
                
                # Check for match of Anlagen data file
                ma = pa.match(line)
                if ma:
                    cost_cntr = ma.group(1).rstrip()
                    if cost_cntr.isdigit():
                        SAPinput.append((
                            'Anlagen', 
                            cost_cntr.strip(), '',  
                            ma.group(4).strip().upper().strip(),
                            '',
                            tl.convertAmount(ma.group(5)),        # Amount.
                            mc.group(2),                          # Date.
                            ma.group(3).strip(),                  # Year.
                            '', '',
                            ma.group(2).strip()))                 # Anl. num.
                        rownum += 1
                        lineCounter += 1
                
                # Check for match of alternative Anlagen data file
                ma = paa.match(line)
                if ma:
                    cost_cntr = ma.group(2).rstrip()
                    if cost_cntr.isdigit():
                        SAPinput.append((
                            'Anlagen', 
                            cost_cntr.strip(), '',
                            ma.group(5).strip().upper().strip(),
                            '',
                            tl.convertAmount(ma.group(6)),        # Amount.
                            ma.group(3),                          # Date.
                            ma.group(4).strip(),                  # Year.
                            '', '',
                            ma.group(1).strip()))                 # Ref. num.
                        rownum += 1
                        lineCounter += 1

                # Check for match of GWG data file
                ma = pg.match(line)
                if ma:
                    cost_cntr = ma.group(1).rstrip()
                    if cost_cntr.isdigit():
                        SAPinput.append((
                            'Anlagen',
                            cost_cntr.strip(), '',
                            ma.group(5).strip().upper().strip(),
                            '',
                            tl.convertAmount(ma.group(4)),        # Amount.
                            ma.group(2),                          # Date.
                            ma.group(3).strip(),                  # Year.
                            '', '',
                            ma.group(8).strip()))                 # Ref. num.
                        rownum += 1
                        lineCounter += 1
            rf.close()
            print("Read " + str(lineCounter) + " of " + \
                    str(totalLineCounter) + " lines from " + str(f))
    
    # Just get a limited number of items in test mode
    SAPinput = SAPinput[:test_len] if test else SAPinput    
    
    #print(SAPinput)
    # Generate a dataframe from input
    df = pd.DataFrame(data=SAPinput, \
                     columns=['Art', 'Kostenstelle', 'Tag', \
                     'Text', 'Kostenart', 'Betrag', 'Datum', 'Jahr', \
                     'Beschaffungsklasse', 'CKA', 'Referenz'])
    
    df['Datum'] = pd.to_datetime(df['Datum'], format='%d.%m.%Y')

    return df

# Writing the results to an Excel file
def writeXLS(filename_w, results):
    from openpyxl import Workbook

    wb = Workbook()
    sheet = wb.active
    
    # Add header line
    sheet['A1'].value = 'Art'
    sheet['B1'].value = 'Kostenstelle'
    sheet['C1'].value = 'System'
    sheet['D1'].value = 'Beschreibung'
    sheet['E1'].value = 'Kostenart'
    sheet['F1'].value = 'Betrag'
    sheet['G1'].value = 'Datum'
    sheet['H1'].value = 'Jahr'
    sheet['I1'].value = 'Referenz'
     
    # Copy data from variable to Sheet
    i = 2
    for wd in results:
        sheet['A' + str(i)].value = wd[3]
        sheet['B' + str(i)].value = wd[4]
        sheet['C' + str(i)].value = ''
        sheet['D' + str(i)].value = wd[2]
        sheet['E' + str(i)].value = wd[5]
        sheet['F' + str(i)].value = wd[7]
        sheet['G' + str(i)].value = wd[8]
        sheet['H' + str(i)].value = wd[9]
        sheet['I' + str(i)].value = wd[12]
        i += 1
        
    # Add Autofilter
    sheet.auto_filter.ref = "A1:H1" + str(i)
    
    # Fixate the top row
    c = sheet['A2']
    sheet.freeze_panes = c

    # Define column width
    sheet.column_dimensions['D'].width = 25
    
    # Don't forget to save the file
    wb.save(filename = filename_w)
    return None

# Coloring a set of cells in a row
def color_row(sheet, cell_range, color):
    from openpyxl.styles import PatternFill
    
    fill = PatternFill("solid", fgColor = color)
    
    # Iterate through the cell range
    rows = sheet[cell_range]
    for row in rows:
        
        # Iterate through the cells
        for c in row:
            c.fill = fill

# Writing the results to an Excel file
def writeXLS(filename_w, results, tag_col, text_col):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import NamedStyle


    wb = Workbook()
    sheet = wb.active

    a = 1            
    for col_name in results.columns:
        if results[col_name].dtype == 'datetime64[ns]':
            sheet.column_dimensions[get_column_letter(a)].width = 10
        if col_name == "Tag":
            col_name = tag_col
            sheet.column_dimensions[get_column_letter(a)].width = 12
        if col_name == "Text":
            col_name = text_col
            sheet.column_dimensions[get_column_letter(a)].width = 60
        sheet.cell(row=1, column=a).value = col_name
        a += 1

    # Start in row 2 after the header row
    i = 2
    for col_name, wd in results.iterrows():
        
        # Assign the rest in a loop
        a = 1
        for cell_value in wd:
            if type(cell_value) == pd._libs.tslibs.timestamps.Timestamp:
                cell_value = cell_value.date()
            sheet.cell(row=i, column=a).value = cell_value
            a += 1

        if i > 1 and type(sheet['M'+str(i)].value) == float:
            # Color the line until quality value with color
            # First get the value and set the color defaults
            red = 'efbdbd'
            yellow = 'efeebd'
            val = float(sheet['M'+str(i)].value)
            
            if val < 0.7:
                color_row(sheet, 'A'+str(i)+':'+'M'+str(i), red)
            if val >= 0.7 and val < 0.9:
                color_row(sheet, 'A'+str(i)+':'+'M'+str(i), yellow)
        i += 1
            
    # Add Autofilter
    sheet.auto_filter.ref = "A1:BA" + str(i)

    # Fixate the top row
    c = sheet['A2']
    sheet.freeze_panes = c

    # Don't forget to save the file
    wb.save(filename = filename_w)
    return None

# Read the excludes file into dataframe 
def readExclude(filename):
    print("Reading Kostenstellen exclude list")
    try:
        df = pd.read_csv(filename, delimiter=' ')
    except FileNotFoundError:
        sys.exit('File ' + filename + ' not found.')
    return df