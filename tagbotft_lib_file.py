import re
import glob
import time
import datetime
import pandas as pd
import numpy as np
import __main__

import tagbotft_lib as tl

# Dickmann / Birkenkamp
# Function to convert a string with numbers in it to a float.
# SAP reports have the minus at the end and put a lot of 
# spaces in it and have a "," instead of a "." as the decimal point.
# In the amount, delete points and trailing whitespace, substitute
# commas by points, and put a possible trailing minus sign at its
# beginning.
def convertAmount(amount: str) -> float:
    temp = amount.strip().replace('.', '').replace(',', '.')
    return float(temp) if temp[-1] != '-' else \
    float('-' + temp[:-1])

# Filter input dataframe by a filter dataframe
def filter_input(in_df, filter_df):
    # Iterate over all columns in the filter dataframe / exclude list
    if len(filter_df) > 0:
        for col in filter_df:
            newData = in_df[~in_df[[col]]\
            .isin(filter_df[col].astype(str).tolist()).any(axis=1)]
        return newData
    else:
        return in_df

# Read learn data from an Excel file
def read_xl_learn_data(f, org_data):

    from openpyxl import load_workbook

    # Loading data file with a maximum number of lines
    tl.message("Reading Excel data sheet: " + f)
    wb = load_workbook(f, data_only=True, read_only=True)
    sheet = wb.active
    
    # Get the real maximum of lines to read
    if sheet.max_row < __main__.max_lines:
         rows = sheet.max_row
    
    # Start with no lines read
    line_count = 0
    
    for row in sheet.rows:
        wsrow = []
        col_i = 0
        for cell in row:
            if col_i < __main__.max_cols:
                wsrow.append(cell.value)
                col_i += 1
        if line_count > __main__.max_lines:
            break
        # Skip if a header row has been already read
        if (len(org_data) > 0) and (line_count == 0):
            line_count += 1
            continue
        else:
            org_data.append(wsrow)
            # Calculating the progress as percentage
            progress = line_count / __main__.max_lines * 100
            line_count += 1
            print("Reading file: %.0f %%" % round(progress), end='\r', flush=True)
    print(f'\n\rLoaded {line_count} lines.')
    wb.close()

    return org_data

# Read learn data from one or multiple Excel files
def read_xl_learn(files):

    # Get the file list
    file_list = list(glob.glob(files))
    org_data = []    

    # Iterate the list of files
    for f in file_list:
        try:
            rf = open(f, encoding='latin-1')
        except FileNotFoundError:
            print('File ' + f + ' not found.')
        else:
            org_data = read_xl_learn_data(f, org_data)

    print("\n\rHeaders:", org_data[0])

    # Arrange the Excel data in a Pandas dataframe
    dfs = pd.DataFrame(data=org_data[1:], columns=org_data[0])

    # Setting default colums in dataframe
    dfs = dfs.rename(columns={__main__.tag_col: 'Tag', \
        __main__.text_col: 'Text'})

    # Ensure consistent column data for columns of string or date value
    for sheet_col in dfs:
        if dfs[sheet_col].dtypes == object:
            dfs[sheet_col] = dfs[sheet_col].astype(str)

    print(f'\n\rLoaded {len(dfs)} lines.')

    return dfs

# Write log file
def writeLog(logEvent):
    f = open('tagbotft.log', 'a')
    dt = datetime.datetime.now()
    stamp = dt.strftime("%Y-%m-%d %H:%M:%S")
    print(logEvent)
    f.write(str(stamp) + " : " + logEvent + "\n")
    f.close()
    
# Dickmann / Birkenkamp
# Read SAP data files with RegEx to identify relevant data lines and items
# Input files can come from Kosten, GWGs oder Anlagen
def read_SAP_1(files: str, exclude_file, work_data):

    # Reading exclude file
    exclude_df = readExclude(exclude_file)

    tl.message("Loading SAP data files")

    # Extract unclassified data from possibly several files, ignoring
    # data rows with cost centres contained in a list of cost centres to
    # be excluded.
    SAPinput = []
    rownum = 0
    
    # Get the list of files to process
    file_list = list(glob.glob(files))
    
    for f in file_list:
        try:
            rf = open(f, encoding='latin-1')
        except FileNotFoundError:
            print('File ' + f + ' not found.')
        else:
            # Counter for all lines
            totalLineCounter = 0
            lineCounter = 0
            # Regular expression pattern for cost files.
            pc = re.compile(
                r'\|(\d{4})\s+'                # Jahr.
                 '\|(\d+\s+)'                  # Kostenstelle.
                 '\|(\d+\s+)'                  # Kostenart.
                 '\|[^\|]+'                    # Kostenartenbezeichn.
                 '\|(\d{2}\.\d{2}\.\d{4})\s+'  # Buchungsdatum.
                 '\|\s*([\d\.]+,\d{2}[\s-])'   # Wert/KWähr.
                 '\|([^\|]+)'                  # Bezeichnung.
                 '\|([^\|]+)'                  # Bestelltext.
                 '\|[^\|]+'                    # Partnerobjektbezeichnung.
                 '\|\s+'                       # Partner-Kstl.
                 '\|(\d+|)\s+'                 # Einkaufsbeleg.
                 '\|\d*\s*'                    # Konto Gegenbuchung.
                 '\|([^\|]+)')               # Ref Beleg.
            # Regular expression pattern for asset files.
            pa = re.compile(
                r'\|([^\|]+)'                  # Kostenst.
                 '\|\d+\s+'                    # Finanzst.
                 '\|\d+\s*'                    # UNr.
                 '\|(\d{2}\.\d{2}\.(\d{4}))'   # Aktivdatum.
                 '\|([^\|]+)'                  # Anlagenbezeichnung.
                 '\|\s*([\d\.]+,\d{2})\s'      # AnschWert.
                 '\|[^\|]+'                    # Währg.
                 '\|[^\|]+'                    # Anlagenbezeichnung.
                 '\|(\d{7})[\s\d]'             # Anlage.
                 '\|\d+\s+\|')                 # Geldg.
            # Regular expression pattern for alternative asset files.
            paa = re.compile(
                r'\|(\d{7})\s*'                # Anlage.
                 '\|\d*\s*'                    # UNr.
                 '\|(\d{7})\s*'                # Kostenst.
                 '\|\d*\s*'                    # Finanzst.
                 '\|(\d{2}\.\d{2}\.(\d{4}))'   # Aktivdatum.
                 '\|([^\|]+)'                  # Anlagenbezeichnung.
                 '\|\s*([\d\.]+,\d{2})\s'      # AnschWert.
                 '\|[^\|]+'                    # Währg.
                 '\|[^\|]+'                    # Anlagenbezeichnung.
                 '\|\d*\s*\|')                 # Geldg.
            # Regular expression pattern for small asset files.
            pg = re.compile(
                r'\|[^\|]+'                    # Kostenst.
                 '\|(\d+\s+)'                  # Finanzst.
                 '\|[^\|]+'                    # Fistl-Text 
                 '\|\d+\s+'                    # Finanzpos.
                 '\|[^\|]+'                    # Vorg. Nr.
                 '\|(\d{2}\.\d{2}\.(\d{4}))'   # HHM-Budat
                 '\|\s*([\d\.]+,\d{2}[\s-])'   # Zahl.Budg.
                 '\|[^\|]+'                    # Kreditor
                 '\|([^\|]+)'                  # Text
                 '\|([^\|]+)'                  # Kurztext
                 '\|([^\|]+)'                  # Bezeichg
                 '\|(\d+\s+|\d+|\s+)'          # RefBelegnr
                 '\|\s+\d+'                    # Pos.
                 '\|\d+\s+'                    # Btr.art
                 '\|[^\|]+'                    # Betragsart
                 '\|[^\|]+'                    # Fipos-Text
                 '\|([^\|]+)')                 # RW-Beleg
            for line in rf:
                totalLineCounter += 1
                # Check for match of Kosten data file
                mc = pc.match(line)
                if mc:
                    cost_cntr = mc.group(2).rstrip()
                    if cost_cntr.isdigit():
                        SAPinput.append((
                            'Kosten', 
                            cost_cntr.strip(), '',                # Kostenart 
                            (mc.group(6).strip() + ' ' + \
                             mc.group(7).strip()).upper(),       # Text.
                            mc.group(3).strip(),
                            convertAmount(mc.group(5)),        # Amount.
                            mc.group(4),                          # Date.
                            mc.group(1).strip(),                  # Year.
                            'b', '',
                            mc.group(8).strip()))                 # Ref. num.
                        rownum += 1
                        lineCounter += 1
                
                # Check for match of Anlagen data file
                ma = pa.match(line)
                if ma:
                    cost_cntr = ma.group(1).rstrip()
                    if cost_cntr.isdigit():
                        SAPinput.append((
                            'Anlagen', 
                            cost_cntr.strip(), '',  
                            ma.group(4).strip().upper().strip(),
                            '',
                            convertAmount(ma.group(5)),        # Amount.
                            mc.group(2),                          # Date.
                            ma.group(3).strip(),                  # Year.
                            'b', '',
                            ma.group(2).strip()))                 # Anl. num.
                        rownum += 1
                        lineCounter += 1
                
                # Check for match of alternative Anlagen data file
                ma = paa.match(line)
                if ma:
                    cost_cntr = ma.group(2).rstrip()
                    if cost_cntr.isdigit():
                        SAPinput.append((
                            'Anlagen', 
                            cost_cntr.strip(), '',
                            ma.group(5).strip().upper().strip(),
                            '',
                            convertAmount(ma.group(6)),        # Amount.
                            ma.group(3),                          # Date.
                            ma.group(4).strip(),                  # Year.
                            'b', '',
                            ma.group(1).strip()))                 # Ref. num.
                        rownum += 1
                        lineCounter += 1

                # Check for match of GWG data file
                ma = pg.match(line)
                if ma:
                    cost_cntr = ma.group(1).rstrip()
                    if cost_cntr.isdigit():
                        SAPinput.append((
                            'Anlagen',
                            cost_cntr.strip(), '',
                            ma.group(5).strip().upper().strip(),
                            '',
                            convertAmount(ma.group(4)),        # Amount.
                            ma.group(2),                          # Date.
                            ma.group(3).strip(),                  # Year.
                            'b', '',
                            ma.group(8).strip()))                 # Ref. num.
                        rownum += 1
                        lineCounter += 1
            rf.close()
            print("Read " + str(lineCounter) + " of " + \
                    str(totalLineCounter) + " lines from " + str(f))
    
    # Just get a limited number of items in test mode
    SAPinput = SAPinput[:__main__.max_input] if __main__.test else SAPinput    
    
    #print(SAPinput)
    # Generate a dataframe from input
    df = pd.DataFrame(data=SAPinput, \
                     columns=['Art', 'Kostenstelle', 'Tag', \
                     'Text', 'Kostenart', 'Betrag', 'Datum', 'Jahr', \
                     'Beschaffungsklasse', 'CKA', 'Referenz'])
    
    df['Datum'] = pd.to_datetime(df['Datum'], format='%d.%m.%Y')
    df = df.replace(r"^None$", '', regex=True)

    # Filtering the the SAP data by the exclude data
    if len(exclude_df) > 1:
        newData = filter_input(df, exclude_df)
    else:
        newData = df 

    # Remove uneven string columns
    if len(work_data) > 1:
        newData = tl.maxcol(newData, work_data)

    return newData

# Coloring a set of cells in a row
def color_row(sheet, cell_range, color):
    from openpyxl.styles import PatternFill
    
    fill = PatternFill("solid", fgColor = color)
    
    # Iterate through the cell range
    rows = sheet[cell_range]
    for row in rows:
        
        # Iterate through the cells
        for c in row:
            c.fill = fill

# Writing the results to an Excel file
def writeXLS(filename_w, results):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import NamedStyle
    from openpyxl.styles import Font, Fill

    wb = Workbook()
    sheet = wb.active
    
    a = 1            
    for col_name in results.columns:
        if results[col_name].dtype == 'datetime64[ns]':
            sheet.column_dimensions[get_column_letter(a)].width = 10
        if col_name == "Tag":
            col_name = __main__.tag_col
            sheet.column_dimensions[get_column_letter(a)].width = 12
        if col_name == "Text":
            col_name = __main__.text_col
            sheet.column_dimensions[get_column_letter(a)].width = 60
        sheet.cell(row=1, column=a).value = col_name
        sheet.cell(row=1, column=a).font = Font(name='Arial', size=10)
        a += 1

    # Start in row 2 after the header row
    i = 2
    for col_name, wd in results.iterrows():
        
        # Assign the rest in a loop
        a = 1
        for cell_value in wd:
            if type(cell_value) == pd._libs.tslibs.timestamps.Timestamp:
                cell_value = cell_value.date()
            sheet.cell(row=i, column=a).value = cell_value
            sheet.cell(row=i, column=a).font = Font(name='Calibri', size=10)
            a += 1

        if i > 1 and type(sheet['M'+str(i)].value) == float:
            # Color the line until quality value with color
            # First get the value and set the color defaults
            red = 'efbdbd'
            yellow = 'efeebd'
            val = float(sheet['M'+str(i)].value)
            
            if val < 0.7:
                color_row(sheet, 'A'+str(i)+':'+'M'+str(i), red)
            if val >= 0.7 and val < 0.9:
                color_row(sheet, 'A'+str(i)+':'+'M'+str(i), yellow)
        i += 1
            
    # Add Autofilter
    sheet.auto_filter.ref = "A1:BA" + str(i)

    # Set auto adjusted column widths
    for idx, col in enumerate(sheet.columns, 1):
        sheet.column_dimensions[get_column_letter(idx)].auto_size = True

    # Fixate the top row
    c = sheet['A2']
    sheet.freeze_panes = c

    # Don't forget to save the file
    wb.save(filename = filename_w)
    return None

# Read the excludes file into dataframe 
def readExclude(filename):
    tl.message("Reading exclude list")
    
    try:
        # Pandas CSV function takes the first line as column name
        df = pd.read_csv(filename, delimiter=' ')
        print("Columns with exclude data:", list(df.columns))
        return df

    # If there is no file return an empty data frame
    except FileNotFoundError:
        print("No exclude file")
        return pd.DataFrame()

def read_CSV(files: str, exclude_file, work_data):

    # Reading exclude file
    exclude_df = readExclude(exclude_file)

    tl.message("Loading CSV data files")

    # Get the list of files to process
    file_list = list(glob.glob(files))
    

    for f in file_list:
        try:
            # Pandas CSV function takes the first line as column name
            if __main__.test is True:
                df = pd.read_csv(f, delimiter=';', nrows=__main__.max_lines, decimal=",")
            else:
                df = pd.read_csv(f, delimiter=';', decimal=",")
        except FileNotFoundError:
            print('File ' + f + ' not found.')
    
    print(len(df),"Entries read...")
    
    df['Date'] = pd.to_datetime(df['Date'], format='%d.%m.%Y')
    df.insert(0,'Tag','')

    # Filtering the the SAP data by the exclude data
    newData = filter_input(df, exclude_df)

    # Remove uneven string columns
    newData = tl.maxcol(newData, work_data)

    return newData
